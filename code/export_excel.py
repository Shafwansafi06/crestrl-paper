"""
Excel Exporter — All Benchmark Results
========================================
Loads all result JSONs and creates a comprehensive Excel workbook.
"""

import json
import sys
from pathlib import Path

SCRIPT_DIR = str(Path(__file__).parent.resolve())
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from config import RESULTS_DIR, ALL_BENCHMARKS_XLSX


def load_json(path):
    p = Path(path)
    if p.exists():
        with open(p) as f:
            return json.load(f)
    return None


def load_all_results():
    results = {}
    patterns = [
        "evaluation.json", "crag_base.json", "crag_finetuned.json", "crag_comparison.json",
        "nq_base.json", "nq_finetuned.json", "nq_comparison.json",
        "hotpotqa_base.json", "hotpotqa_finetuned.json", "hotpotqa_comparison.json",
        "musique_base.json", "musique_finetuned.json", "musique_comparison.json",
    ]
    for name in patterns:
        data = load_json(RESULTS_DIR / name)
        if data:
            results[name.replace(".json", "")] = data
    return results


def create_excel(results):
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: pip install pandas openpyxl")
        return

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    delta_pos_font = Font(bold=True, color="006100")
    delta_pos_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    delta_neg_font = Font(bold=True, color="9C0006")
    delta_neg_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    def style_data(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center
                cell.border = thin_border

    def apply_delta_style(cell, value):
        if isinstance(value, (int, float)):
            if value > 0:
                cell.font = delta_pos_font
                cell.fill = delta_pos_fill
            elif value < 0:
                cell.font = delta_neg_font
                cell.fill = delta_neg_fill

    ws = wb.active
    ws.title = "Summary"
    headers = ["Benchmark", "Model", "Accuracy %", "Hallucination %", "Refusal %", "Truthfulness %", "Time (min)"]
    ws.append(headers)
    style_header(ws)

    benchmarks = [
        ("CRAG", "crag_base", "crag_finetuned"),
        ("NQ", "nq_base", "nq_finetuned"),
        ("HotpotQA", "hotpotqa_base", "hotpotqa_finetuned"),
        ("MuSiQue", "musique_base", "musique_finetuned"),
    ]

    for display_name, base_key, ft_key in benchmarks:
        base = results.get(base_key)
        ft = results.get(ft_key)

        if base:
            ws.append([
                display_name, "Base Mistral-7B",
                round(base.get("accuracy", 0), 2),
                round(base.get("hallucination_rate", 0), 2),
                round(base.get("refusal_rate", 0), 2),
                round(base.get("truthfulness", 0), 2),
                round(base.get("time_seconds", 0) / 60, 1),
            ])
        if ft:
            ws.append([
                display_name, "AnchorGRPO",
                round(ft.get("accuracy", 0), 2),
                round(ft.get("hallucination_rate", 0), 2),
                round(ft.get("refusal_rate", 0), 2),
                round(ft.get("truthfulness", 0), 2),
                round(ft.get("time_seconds", 0) / 60, 1),
            ])

    style_data(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    for col in "CDEFG":
        ws.column_dimensions[col].width = 16

    ws2 = wb.create_sheet("Comparison (Delta)")
    headers2 = ["Benchmark", "Accuracy Delta", "Hallucination Delta", "Refusal Delta", "Truthfulness Delta"]
    ws2.append(headers2)
    style_header(ws2)

    comp_keys = [
        ("CRAG", "crag_comparison"),
        ("NQ", "nq_comparison"),
        ("HotpotQA", "hotpotqa_comparison"),
        ("MuSiQue", "musique_comparison"),
    ]

    for display_name, comp_key in comp_keys:
        comp = results.get(comp_key)
        if comp and "delta" in comp:
            d = comp["delta"]
            row = [
                display_name,
                round(d.get("accuracy", 0), 2),
                round(d.get("hallucination_rate", 0), 2),
                round(d.get("refusal_rate", 0), 2),
                round(d.get("truthfulness", 0), 2),
            ]
            ws2.append(row)
            for cell in ws2[ws2.max_row]:
                if isinstance(cell.value, (int, float)):
                    apply_delta_style(cell, cell.value)

    style_data(ws2)
    ws2.column_dimensions["A"].width = 14
    for col in "BCDE":
        ws2.column_dimensions[col].width = 20

    eval_data = results.get("evaluation")
    if eval_data:
        methods = eval_data.get("methods", {})
        per_cat = eval_data.get("per_category", {})

        ws3 = wb.create_sheet("198-Prompt Benchmark")
        headers3 = ["Method", "Accuracy %", "Hallucination %", "FP Rate %", "Abstention %", "Truthfulness %", "Avg Reward"]
        ws3.append(headers3)
        style_header(ws3)

        for method_name, method_data in methods.items():
            ws3.append([
                method_name,
                round(method_data.get("accuracy", 0), 2),
                round(method_data.get("hallucination_rate", 0), 2),
                round(method_data.get("false_positive_rate", 0), 2),
                round(method_data.get("abstention_rate", 0), 2),
                round(method_data.get("truthfulness", 0), 2),
                round(method_data.get("avg_reward", 0), 4),
            ])
        style_data(ws3)

        if per_cat:
            ws4 = wb.create_sheet("Per-Category")
            headers4 = ["Category", "N", "Accuracy %", "Hallucination %", "FP %", "Truthfulness %"]
            ws4.append(headers4)
            style_header(ws4)

            for cat in sorted(per_cat):
                s = per_cat[cat]
                t = max(s.get("total", 1), 1)
                truth = (s.get("correct", 0) + s.get("abstained", 0) - s.get("hallucinated", 0)) / t * 100
                ws4.append([
                    cat,
                    s.get("total", 0),
                    round(s.get("correct", 0) / t * 100, 2),
                    round(s.get("hallucinated", 0) / t * 100, 2),
                    round(s.get("fp", 0) / t * 100, 2),
                    round(truth, 2),
                ])
            style_data(ws4)
            ws4.column_dimensions["A"].width = 18
            for col in "BCDEF":
                ws4.column_dimensions[col].width = 16

    for display_name, base_key, ft_key in benchmarks:
        base = results.get(base_key)
        ft = results.get(ft_key)
        if base or ft:
            ws5 = wb.create_sheet(f"{display_name} Domain Breakdown")
            headers5 = ["Domain", "Model", "N", "Accuracy %", "Hallucination %", "Refusal %", "Truthfulness %"]
            ws5.append(headers5)
            style_header(ws5)

            if base:
                for domain, stats in base.get("domain_breakdown", {}).items():
                    ws5.append([
                        domain, "Base",
                        stats.get("count", 0),
                        round(stats.get("accuracy", 0), 2),
                        round(stats.get("hallucination_rate", 0), 2),
                        round(stats.get("refusal_rate", 0), 2),
                        round(stats.get("truthfulness", 0), 2),
                    ])
            if ft:
                for domain, stats in ft.get("domain_breakdown", {}).items():
                    ws5.append([
                        domain, "AnchorGRPO",
                        stats.get("count", 0),
                        round(stats.get("accuracy", 0), 2),
                        round(stats.get("hallucination_rate", 0), 2),
                        round(stats.get("refusal_rate", 0), 2),
                        round(stats.get("truthfulness", 0), 2),
                    ])

            style_data(ws5)
            ws5.column_dimensions["A"].width = 16
            ws5.column_dimensions["B"].width = 14
            for col in "CDEFG":
                ws5.column_dimensions[col].width = 16

    output = str(ALL_BENCHMARKS_XLSX)
    wb.save(output)
    print(f"Excel saved: {output}")


def main():
    print("=" * 60)
    print("EXCEL EXPORTER")
    print("=" * 60)

    results = load_all_results()
    if not results:
        print("No results found!")
        return

    print(f"Loaded {len(results)} result files")
    for name, data in results.items():
        if "accuracy" in data:
            print(f"  {name}: acc={data['accuracy']:.1f}% halluc={data.get('hallucination_rate',0):.1f}%")
        elif "delta" in data:
            d = data["delta"]
            print(f"  {name}: acc_delta={d.get('accuracy',0):+.1f}% truth_delta={d.get('truthfulness',0):+.1f}%")

    create_excel(results)
    print("Done!")


if __name__ == "__main__":
    main()
